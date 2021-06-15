from django.shortcuts import render, redirect, reverse
from django.http import HttpResponse
from django.views import generic
from django.contrib.auth.mixins import LoginRequiredMixin
from .models import Contact, User
from .forms import ContactForm, ContactModelForm, CustomUserCreationForm
import xlwt
import openpyxl
import sqlite3


class SignupView(generic.CreateView):
  template_name = "registration/signup.html"
  form_class = CustomUserCreationForm
  def get_success_url(self):
    return reverse("login")

class LandingPageView(generic.TemplateView):
  template_name = "landing_page.html"

class ContactListView(LoginRequiredMixin, generic.ListView):
  template_name = "contact_list.html"
  context_object_name = 'contacts'

  def get_queryset(self):
    queryset = Contact.objects.all()
    if self.request.user.is_authenticated:
      queryset = queryset.filter(user=self.request.user)
    return queryset

class ContactDetailView(LoginRequiredMixin, generic.DetailView):
  template_name = "contact_detail.html"
  context_object_name = 'contact'

  def get_queryset(self):
    queryset = Contact.objects.all()
    if self.request.user.is_authenticated:
      queryset = queryset.filter(user=self.request.user)
    return queryset

class ContactCreateView(LoginRequiredMixin, generic.CreateView):
  template_name = "contact_create.html"
  form_class = ContactModelForm

  def get_success_url(self):
    return reverse("contacts:contact-list")

  def form_valid(self, form):
    contact = form.save(commit=False)
    contact.user = self.request.user
    contact.save()
    return super(ContactCreateView, self).form_valid(form)

class ContactUpdateView(LoginRequiredMixin, generic.UpdateView):
  template_name = "contact_update.html"
  form_class = ContactModelForm

  def get_queryset(self):
    queryset = Contact.objects.all()
    if self.request.user.is_authenticated:
      queryset = queryset.filter(user=self.request.user)
    return queryset

  def get_success_url(self):
    return reverse("contacts:contact-list")

class ContactDeleteView(LoginRequiredMixin, generic.DeleteView):
  template_name = "contact_delete.html"
  form_class = ContactModelForm

  def get_queryset(self):
    queryset = Contact.objects.all()
    if self.request.user.is_authenticated:
      queryset = queryset.filter(user=self.request.user)
    return queryset

  def get_success_url(self):
    return reverse("contacts:contact-list")

def export_contacts_xls(request):
  response = HttpResponse(content_type='application/ms-excel')
  response['Content-Disposition'] = 'attachment; filename="users.xls"'

  wb = xlwt.Workbook(encoding='utf-8')
  ws = wb.add_sheet("Contact List")

  row_num = 0
  font_style = xlwt.XFStyle()
  font_style.font.bold = True

  columns = ['first_name', 'last_name', 'email', 'address_1', 'address_2', 'city', 'state', 'zipcode', 'profile_photo', 'user']

  for col_num in range(len(columns)):
    ws.write(row_num, col_num, columns[col_num], font_style)

  font_style = xlwt.XFStyle()

  rows = Contact.objects.filter(user=request.user).values_list('first_name', 'last_name', 'email', 'address_1', 'address_2', 'city', 'state', 'zipcode', 'profile_photo', 'user')

  for row in rows:
    row_num += 1
    for col_num in range(len(row)):
      ws.write(row_num, col_num, row[col_num], font_style)

  wb.save(response)
  return response

# step 1 - create a view to handle uploading xls file
# step 2 - load xls file with openpyxl
# step 3 - seed db with xls data
# step 4 - render updated contact list on html

def import_contacts_xls(request):
  if request.method == "GET":
    return render(request, 'contact_upload.html', {})
  else:
    excel_file = request.FILES['excel_file']

    wb = openpyxl.load_workbook(excel_file)
    active_sheet = wb.active

    current_user = request.user

    for row in active_sheet.iter_rows(min_row=2):
      Contact.objects.create(user=current_user, first_name=row[0].value, last_name=row[1].value, email=row[2].value, address_1=row[3].value, address_2=row[4].value, city=row[5].value, state=row[6].value, zipcode=row[7].value)

    response = 'Upload Successful!'

  return render(request, 'contact_upload.html', {'success':response})

